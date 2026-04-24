"""
Export confusion matrix values (TP, TN, FP, FN) for all 125 combinations to Excel.
One summary sheet + one sheet per dataset with color-coded cells.
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
import os

df = pd.read_csv("updated_model_results.csv")

TEST_INFO = {
    'D1': {'N': 1_272_524, 'P':     1_643, 'name': 'Dataset 1'},
    'D2': {'N':     2_000, 'P':        30, 'name': 'Dataset 2'},
    'D3': {'N':   113_726, 'P':    56_863, 'name': 'Dataset 3'},
    'D4': {'N':    56_962, 'P':        98, 'name': 'Dataset 4'},
    'D5': {'N':   259_335, 'P':     1_501, 'name': 'Dataset 5'},
}

PRIM = {'M1':'HistGradBoost','M2':'ExtraTrees','M3':'GradBoost',
        'M4':'RandomForest','M5':'MLP'}
SEC  = {'M6':'Isolation Forest','M7':'Autoencoder',
        'M8':'VAE','M9':'DAGMM','M10':'Deep SVDD'}

def reconstruct_cm(precision, recall, accuracy, N, P):
    precision = max(float(precision), 1e-9)
    recall    = max(float(recall),    1e-9)
    TP = round(P * recall)
    FN = P - TP
    FP = round(TP * (1 / precision - 1)) if precision < 1.0 else 0
    TN = max(N - TP - FP - FN, 0)
    return int(TP), int(TN), int(FP), int(FN)

# ── Build full table ──────────────────────────────────────────────────────────
rows = []
for _, r in df.iterrows():
    info = TEST_INFO[r['Dataset']]
    TP, TN, FP, FN = reconstruct_cm(r['Precision'], r['Recall'],
                                     r['Accuracy'],  info['N'], info['P'])
    spec = TN / (TN + FP) if (TN + FP) > 0 else 0
    rows.append({
        'Dataset':            r['Dataset'],
        'Dataset Name':       info['name'],
        'Primary Model':      f"{r['Primary_Model']} ({PRIM.get(r['Primary_Model'], '')})",
        'Secondary Model':    f"{r['Secondary_Model']} ({SEC.get(r['Secondary_Model'], '')})",
        'Hybrid Combination': r['Hybrid_Combination'],
        'TP (True Positive)':  TP,
        'TN (True Negative)':  TN,
        'FP (False Positive)': FP,
        'FN (False Negative)': FN,
        'Accuracy':           round(r['Accuracy'],  6),
        'Precision':          round(r['Precision'], 6),
        'Recall (Sensitivity)': round(r['Recall'],  6),
        'Specificity':        round(spec, 6),
        'F1 Score':           round(r['F1_Score'],  6),
        'AUC':                round(r['AUC'],        6),
        'Time (s)':           round(r['Time_Seconds'], 2),
    })

full_df = pd.DataFrame(rows)

# ── Colour helpers ────────────────────────────────────────────────────────────
def hex_fill(hex_str):
    return PatternFill(start_color=hex_str, end_color=hex_str, fill_type='solid')

def metric_fill(val, low=0.5, high=1.0):
    """Green gradient based on value."""
    t = max(0, min(1, (val - low) / (high - low)))
    r  = int(255 * (1 - t))
    g  = int(200 * t + 55)
    b  = 60
    return PatternFill(start_color=f'{r:02X}{g:02X}{b:02X}',
                       end_color=f'{r:02X}{g:02X}{b:02X}', fill_type='solid')

HDR_FILL   = hex_fill('1E3A5F')
BAND_FILLS = [hex_fill('FFFFFF'), hex_fill('EEF4FB')]
GOLD_FILL  = hex_fill('FFF3CD')
thin = Side(style='thin', color='BBBBBB')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

hdr_font  = Font(bold=True, color='FFFFFF', size=10, name='Calibri')
body_font = Font(size=9,  name='Calibri')
title_font= Font(bold=True, size=13, name='Calibri', color='1E3A5F')
center    = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_al   = Alignment(horizontal='left',   vertical='center')

def write_header(ws, cols, row=1):
    for ci, col in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=col)
        c.fill      = HDR_FILL
        c.font      = hdr_font
        c.alignment = center
        c.border    = border

def write_row(ws, values, row, band, metric_cols=None, col_names=None):
    fill = BAND_FILLS[band % 2]
    for ci, val in enumerate(values, 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.border    = border
        c.alignment = center if isinstance(val, (int, float)) else left_al
        c.font      = body_font
        if metric_cols and col_names and col_names[ci-1] in metric_cols and isinstance(val, float):
            c.fill = metric_fill(val)
        else:
            c.fill = fill

# ══════════════════════════════════════════════════════════════════════════════
wb = Workbook()

# ── Sheet 1: Summary (all 125 rows) ──────────────────────────────────────────
ws_all = wb.active
ws_all.title = "All Combinations"
ws_all.freeze_panes = 'A2'

cols = list(full_df.columns)
metric_c = {'Accuracy','Precision','Recall (Sensitivity)','Specificity','F1 Score','AUC'}

# Title row
ws_all.merge_cells('A1:P1')
t = ws_all['A1']
t.value     = '🔍 Hybrid Fraud Detection — Confusion Matrix Results (All 125 Combinations)'
t.font      = title_font
t.alignment = center
t.fill      = hex_fill('D6E4F0')

write_header(ws_all, cols, row=2)

for i, (_, r) in enumerate(full_df.iterrows()):
    band = i % 2
    vals = [r[c] for c in cols]
    write_row(ws_all, vals, i + 3, band, metric_c, cols)
    # Highlight best per dataset in gold
    ds_best = full_df[full_df['Dataset'] == r['Dataset']]['AUC'].idxmax()
    if _ == ds_best:
        for ci in range(1, len(cols)+1):
            ws_all.cell(row=i+3, column=ci).fill = GOLD_FILL

# Column widths
widths = [8, 12, 22, 22, 18, 14, 14, 14, 14, 11, 11, 18, 11, 11, 10, 10]
for ci, w in enumerate(widths, 1):
    ws_all.column_dimensions[get_column_letter(ci)].width = w
ws_all.row_dimensions[1].height = 22
ws_all.row_dimensions[2].height = 35

# ── Sheet 2+: One per dataset ─────────────────────────────────────────────────
for ds, info in TEST_INFO.items():
    ws = wb.create_sheet(title=info['name'])
    ws.freeze_panes = 'A3'

    sub = full_df[full_df['Dataset'] == ds].reset_index(drop=True)

    # Title
    ws.merge_cells('A1:P1')
    t = ws['A1']
    t.value     = f"📊 {info['name']} — Confusion Matrix for All 25 Hybrid Combinations"
    t.font      = title_font
    t.alignment = center
    t.fill      = hex_fill('D6E4F0')

    write_header(ws, cols, row=2)

    best_idx = sub['AUC'].idxmax()
    for i, (_, r) in enumerate(sub.iterrows()):
        vals = [r[c] for c in cols]
        write_row(ws, vals, i + 3, i % 2, metric_c, cols)
        if i == best_idx:
            for ci in range(1, len(cols)+1):
                ws.cell(row=i+3, column=ci).fill = GOLD_FILL

    # ── Embedded 2×2 CM for best model ──────────────────────────────────────
    best = sub.loc[best_idx]
    TP = best['TP (True Positive)'];  TN = best['TN (True Negative)']
    FP = best['FP (False Positive)']; FN = best['FN (False Negative)']

    start_row = len(sub) + 5
    ws.merge_cells(f'A{start_row}:H{start_row}')
    hdr = ws[f'A{start_row}']
    hdr.value     = f"⭐ Best Combination: {best['Hybrid Combination']}  |  AUC={best['AUC']:.4f}  F1={best['F1 Score']:.4f}"
    hdr.font      = Font(bold=True, size=11, name='Calibri', color='1E3A5F')
    hdr.alignment = center
    hdr.fill      = hex_fill('FFF3CD')

    cm_data = [
        ['',            'Pred: Legit',     'Pred: Fraud'],
        ['Actual: Legit', int(TN),          int(FP)],
        ['Actual: Fraud', int(FN),          int(TP)],
    ]
    cm_fills = {
        (0, 0): '4CAF50',  # TN green
        (0, 1): 'F44336',  # FP red
        (1, 0): 'FF9800',  # FN orange
        (1, 1): '2196F3',  # TP blue
    }
    for ri, row_data in enumerate(cm_data):
        for ci, val in enumerate(row_data):
            c = ws.cell(row=start_row + 1 + ri, column=1 + ci, value=val)
            c.border    = border
            c.alignment = center
            is_value_cell = (ri > 0 and ci > 0)
            c.font = Font(bold=True, size=10, name='Calibri',
                          color='FFFFFF' if is_value_cell else '1E3A5F')
            if is_value_cell:
                c.fill = hex_fill(cm_fills[(ri - 1, ci - 1)])
            else:
                c.fill = hex_fill('D6E4F0')

    for ci in range(1, len(cols)+1):
        ws.column_dimensions[get_column_letter(ci)].width = widths[ci-1]
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 35

# ── Sheet: CM Summary table only ─────────────────────────────────────────────
ws_cm = wb.create_sheet(title="CM Summary")
ws_cm.merge_cells('A1:I1')
t = ws_cm['A1']
t.value = '🧮 Confusion Matrix Values Summary — Best Model per Dataset'
t.font = title_font; t.alignment = center; t.fill = hex_fill('D6E4F0')

cm_cols = ['Dataset','Best Combination','TP','TN','FP','FN',
           'Sensitivity (Recall)','Specificity','AUC']
write_header(ws_cm, cm_cols, row=2)

for i, (ds, info) in enumerate(TEST_INFO.items()):
    sub  = full_df[full_df['Dataset'] == ds]
    best = sub.loc[sub['AUC'].idxmax()]
    vals = [
        ds,
        best['Hybrid Combination'],
        int(best['TP (True Positive)']),
        int(best['TN (True Negative)']),
        int(best['FP (False Positive)']),
        int(best['FN (False Negative)']),
        round(best['Recall (Sensitivity)'], 4),
        round(best['Specificity'], 4),
        round(best['AUC'], 4),
    ]
    write_row(ws_cm, vals, i + 3, i % 2,
              {'Sensitivity (Recall)','Specificity','AUC'}, cm_cols)

for ci, w in enumerate([8,25,14,14,14,14,20,14,10], 1):
    ws_cm.column_dimensions[get_column_letter(ci)].width = w
ws_cm.row_dimensions[2].height = 35

out = 'updated_model_results.xlsx'
wb.save(out)
print(f"✅ Excel saved: {out}")
print(f"   Sheets: All Combinations | D1–D5 detail | CM Summary")
print(f"   Rows: {len(full_df)} combinations + embedded 2×2 CM per dataset sheet")
print(f"   Gold rows = best AUC per dataset")
